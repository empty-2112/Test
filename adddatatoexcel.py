def exportdata(value):

  import pandas
  import mysql.connector
  import numpy
  import datetime
  from openpyxl import load_workbook
  from openpyxl.utils.dataframe import dataframe_to_rows
  from openpyxl.styles.borders import Border, Side
  
  
  thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

  mydb = mysql.connector.connect(host="113.176.99.177", user="coca_remote", passwd="coca20@123", database="loragateway")

  #x = datetime.datetime.now() 
  x = value
  today = str(x.date())
  Last = []
  Last.append(x)

  for x in range(5):
    x = str(x + 1)
    mycursor = mydb.cursor()
   
    mycursor.execute("SELECT di1 FROM `loragateway`.`node" + x + "_data` WHERE TIMESTAMP LIKE '" + today + "%'")  
    Totalizer_arr = mycursor.fetchall()    
    Totalizer_arr = numpy.array(Totalizer_arr, dtype='f')       
    Totalizer = Totalizer_arr[len(Totalizer_arr) - 1] - Totalizer_arr[0]    
    Last.append(numpy.round(Totalizer[0], decimals = 2))    
  

    mycursor.execute("SELECT AVG(ai2) FROM `loragateway`.`node" + x + "_data` WHERE TIMESTAMP LIKE '" + today + "%' AND ai1 < 1.00 AND ai1 != 'ERR' AND ai2 != 'ERR'")  
    Level_Low_array = mycursor.fetchone()
    Level_Low_array = numpy.array(Level_Low_array, dtype='f') 
    Level_Low =  Level_Low_array[0]
    Last.append(numpy.round(Level_Low, decimals = 2))

    mycursor.execute("SELECT AVG(ai2) FROM `loragateway`.`node" + x + "_data` WHERE TIMESTAMP LIKE '" + today + "%' AND ai1 > 1.00 AND ai1 != 'ERR' AND ai2 != 'ERR'")  
    Level_High_array = mycursor.fetchone()
    Level_High_array = numpy.array(Level_High_array, dtype='f') 
    Level_High =  Level_High_array[0]
    Last.append(numpy.round(Level_High, decimals = 2))

    #mycursor.close()

  
  Last = pandas.DataFrame(Last)
  Last = Last.transpose()
  print(Last)

  wb = load_workbook("Book1.xlsx")

  # Select First Worksheet
  ws = wb.worksheets[0]

  for r in dataframe_to_rows(Last, index=False, header=False):
      ws.append(r)  

  for i in range(1,17):    
    ws.cell(row = ws.max_row, column = i).border = thin_border
    ws.cell(row = ws.max_row, column = i + 1).number_format = '#,##0.00'

  wb.save("Book1.xlsx")
  #mydb.close()  

  
